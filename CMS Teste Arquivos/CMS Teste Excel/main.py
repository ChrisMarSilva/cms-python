from loguru import logger
import datetime as dt
import time
from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font
from openpyxl.drawing.image import Image
# from openpyxl import *
from dotenv import load_dotenv


def criar_excel():

        wb = Workbook()

        # grab the active worksheet
        ws = wb.active

        # Data can be assigned directly to cells
        ws['A1'] = 42

        # Rows can also be appended
        ws.append([1, 2, 3])

        # Python types will automatically be converted
        ws['A2'] = dt.datetime.now()

        # Save the file
        wb.save("sample.xlsx")

        data = {
            "Joe": {
                "math": 65,
                "science": 78,
                "english": 98,
                "gym": 89
            },
            "Bill": {
                "math": 55,
                "science": 72,
                "english": 87,
                "gym": 95
            },
            "Tim": {
                "math": 100,
                "science": 45,
                "english": 75,
                "gym": 92
            },
            "Sally": {
                "math": 30,
                "science": 25,
                "english": 45,
                "gym": 100
            },
            "Jane": {
                "math": 100,
                "science": 100,
                "english": 100,
                "gym": 60
            }
        }

        wb = Workbook()
        ws = wb.active
        ws.title = "Grades"


        headings = ['Name'] + list(data['Joe'].keys())
        ws.append(headings)

        for person in data:
            grades = list(data[person].values())
            ws.append([person] + grades)

        for col in range(2, len(data['Joe']) + 2):
            char = get_column_letter(col)
            ws[char + "7"] = f"=SUM({char + '2'}:{char + '6'})/{len(data)}"

        for col in range(1, 6):
            ws[get_column_letter(col) + '1'].font = Font(bold=True, color="0099CCFF")

        # ws.merge_cells('A2:D2')
        # ws.unmerge_cells('A2:D2')

        # ws.merge_cells(start_row=2, start_column=1, end_row=4, end_column=4)
        # ws.unmerge_cells(start_row=2, start_column=1, end_row=4, end_column=4)

        # img = Image('logo.png')
        # ws.add_image(img, 'A1')
        
        # ws = wb.create_sheet()
        # ws.column_dimensions.group('A','D', hidden=True)
        # ws.row_dimensions.group(1,10, hidden=True)
        
        # sheet['C2'] = '=AVERAGE(A1:A5)'

        wb.save("NewGrades.xlsx")
        # wb.save(filename="./NewGrades.xlsx")

def ler_excel():
    
    # workbook = load_workbook(filename="./sample.xlsx")
    # sheet = workbook['Sheet']

    workbook = load_workbook(filename="./NewGrades.xlsx") # wb = load_workbook(filename="./NewGrades.xlsx")
    sheet = workbook.active # sh = workbook.active
    # sheet = workbook['Grades']

    logger.info(f'row') 
    for row in sheet.iter_rows(values_only=True):
        logger.info(f'{row=}') # (<Cell 'Grades'.A1>, <Cell 'Grades'.B1>, <Cell 'Grades'.C1>, <Cell 'Grades'.D1>, <Cell 'Grades'.E1>)

    logger.info(f'cols') 
    for cols in sheet.iter_cols(values_only=True):
        logger.info(f'{cols=}') # cols=(<Cell 'Grades'.A1>, <Cell 'Grades'.A2>, <Cell 'Grades'.A3>, <Cell 'Grades'.A4>, <Cell 'Grades'.A5>, <Cell 'Grades'.A6>, <Cell 'Grades'.A7>)

    logger.info(f'sheet') 
    logger.info(f'{sheet["A1"].value=}') 
    logger.info(f'{sheet["A1"].number_format=}') 
    logger.info(f'{sheet.cell(row=1, column=1).value=}') 
    logger.info(f'{sheet.cell(row=1, column=1, value=0).value=}') 
    logger.info(f'{sheet["A1:C2"]=}') 
    logger.info(f'{workbook.sheetnames=}') 


def main():
    try:

        logger.info(f'Inicio') 
        start_time = time.perf_counter()  # time.time()  # time.perf_counter()  # time.perf_counter_ns()  # time.process_time()

        # criar_excel()
        ler_excel()
        
        # python main.py

        end_time = time.perf_counter() - start_time  # time.time() # time.perf_counter() # time.perf_counter_ns() # time.process_time()
        logger.info(f"Fim - Done in {end_time:.2f}s - {dt.timedelta(seconds=end_time)}")

    except KeyboardInterrupt:
        pass
    except Exception as e:
        logger.error(f'Falha Geral(main): "{str(e)}"')


if __name__ == '__main__':
    main()

# py -3 -m venv .venv

# python -m pip install --upgrade openpyxl

# cd c:/Users/chris/Desktop/CMS Python/CMS Teste HelloWorld
# .venv\scripts\activate

# python main.py
# python3 main.py  
# pypy3 main.py
